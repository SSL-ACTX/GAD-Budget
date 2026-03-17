import os
import sqlite3
import requests
import openpyxl
from datetime import datetime, date

class SpreadsheetSync:
    def __init__(self, db_path, excel_path=None, sheet_url=None):
        self.db_path = db_path
        self.excel_path = excel_path
        self.sheet_url = sheet_url

    def _connect(self):
        return sqlite3.connect(self.db_path)

    def _safe_float(self, v):
        if v is None or v == "" or v == "-":
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace(",", "")
        try:
            return float(s)
        except ValueError:
            return 0.0

    def sync(self):
        if self.excel_path and os.path.exists(self.excel_path):
            return self.sync_local()
        elif self.sheet_url:
            return self.sync_remote()
        else:
            return False, "No valid data source provided."

    def sync_local(self):
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            
            # 1. MONITORING (budgets table)
            if "MONITORING" in wb.sheetnames:
                self._sync_monitoring(wb["MONITORING"])
            
            # 2. SUMMARY (grand_totals and office_summary)
            if "SUMMARY" in wb.sheetnames:
                self._sync_summary(wb["SUMMARY"])
            
            # 3. LBP2-Accounts (expenditures)
            if "LBP2-Accounts" in wb.sheetnames:
                self._sync_expenditures(wb["LBP2-Accounts"])
            
            # 4. LBP4-AIP-Obligated (aip_programs)
            if "LBP4-AIP-Obligated" in wb.sheetnames:
                self._sync_aip_programs(wb["LBP4-AIP-Obligated"])
                
            return True, "Sync successful from local file."
        except Exception as e:
            return False, f"Error during local sync: {str(e)}"

    def _sync_monitoring(self, sheet):
        data = []
        header_found = False
        for row in sheet.iter_rows(values_only=True):
            if not header_found:
                if "NO." in [str(cell).upper() for cell in row if cell]:
                    header_found = True
                continue
            
            if not any(row): continue
            
            # Map index to column name
            # NO. (0), DATE (1), OFFICE (2), STATUS (3), PARTICULARS (4), 
            # POW TITLE (5), POW DATE (6), REF CODE (7), ALLOTED (8), PPA DESC (9),
            # ACCOUNTS (10), ACC CODE (11), PROPOSED (12), OBLIGATION (13),
            # PAYEE (14), CAF (15), OBR NO (16), TYPE (17), REMAINING (18), REMARKS (19)
            
            # Date handling
            dt = row[1]
            if isinstance(dt, datetime):
                dt_str = dt.strftime("%m/%d/%Y")
            elif isinstance(dt, (int, float)): # Excel serial date
                # Simple conversion for demo
                dt_str = str(dt) 
            else:
                dt_str = str(dt) if dt else ""

            row_dict = {
                "no": str(row[0]) if row[0] is not None else "",
                "date": dt_str,
                "office": str(row[2]) if row[2] else "",
                "status": str(row[3]) if row[3] else "Planned",
                "particulars": str(row[4]) if row[4] else "",
                "pow_title": str(row[5]) if row[5] else "",
                "pow_date_of_activity": str(row[6]) if row[6] else "",
                "reference_code": str(row[7]) if row[7] else "",
                "ppa_allotted_budget": self._safe_float(row[8]),
                "ppa_description": str(row[9]) if row[9] else "",
                "accounts": str(row[10]) if row[10] else "",
                "account_code": str(row[11]) if row[11] else "",
                "proposed_budget": self._safe_float(row[12]),
                "actual_obligation": self._safe_float(row[13]),
                "payee": str(row[14]) if row[14] else "",
                "caf": str(row[15]) if row[15] else "",
                "obligation_number": str(row[16]) if row[16] else "",
                "venue_food_honorarium": str(row[17]) if row[17] else "",
                "actual_remaining_budget": self._safe_float(row[18]),
                "remarks": str(row[19]) if row[19] else "",
                "created_at": datetime.utcnow().isoformat(timespec="seconds")
            }
            data.append(row_dict)

        with self._connect() as conn:
            conn.execute("DELETE FROM budgets")
            conn.executemany("""
                INSERT INTO budgets (
                    no, date, office, status, particulars, pow_title, pow_date_of_activity,
                    reference_code, ppa_allotted_budget, ppa_description, accounts,
                    account_code, proposed_budget, actual_obligation, payee, caf,
                    obligation_number, venue_food_honorarium, actual_remaining_budget,
                    remarks, created_at
                ) VALUES (
                    :no, :date, :office, :status, :particulars, :pow_title, :pow_date_of_activity,
                    :reference_code, :ppa_allotted_budget, :ppa_description, :accounts,
                    :account_code, :proposed_budget, :actual_obligation, :payee, :caf,
                    :obligation_number, :venue_food_honorarium, :actual_remaining_budget,
                    :remarks, :created_at
                )
            """, data)
            conn.commit()

    def _sync_summary(self, sheet):
        rows = list(sheet.iter_rows(values_only=True))
        
        # Grand Totals
        threshold_row = None
        expenses_row = None
        for i, row in enumerate(rows):
            row_str = " ".join([str(c) for c in row if c])
            if "TOTAL GAD BUDGET THRESHOLD" in row_str:
                threshold_row = i + 1
            if "PERSONAL SERVICES" in row_str and "Balance" in row_str:
                expenses_row = i + 1
        
        if threshold_row and threshold_row < len(rows):
            gt_data = rows[threshold_row]
        else:
            gt_data = [0] * 10

        if expenses_row and expenses_row < len(rows):
            ps_co_data = rows[expenses_row]
        else:
            ps_co_data = [0] * 10
        
        grand = {
            "total_gad_threshold": self._safe_float(gt_data[0]),
            "total_obligated": self._safe_float(gt_data[2]),
            "total_earnmarked": self._safe_float(gt_data[3]),
            "total_expenses": self._safe_float(gt_data[4]),
            "utilization_pct": self._safe_float(gt_data[6]),
            "ps_expenses": self._safe_float(ps_co_data[0]),
            "ps_balance": self._safe_float(ps_co_data[2]),
            "mooe_expenses": self._safe_float(ps_co_data[3]),
            "mooe_balance": self._safe_float(ps_co_data[4]),
            "co_expenses": self._safe_float(ps_co_data[5]),
            "co_balance": self._safe_float(ps_co_data[6])
        }

        with self._connect() as conn:
            conn.execute("DELETE FROM grand_totals")
            conn.execute("""
                INSERT INTO grand_totals (
                    total_gad_threshold, total_obligated, total_earnmarked, total_expenses,
                    utilization_pct, ps_expenses, ps_balance, mooe_expenses, mooe_balance,
                    co_expenses, co_balance
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, (
                grand["total_gad_threshold"], grand["total_obligated"], grand["total_earnmarked"],
                grand["total_expenses"], grand["utilization_pct"], grand["ps_expenses"],
                grand["ps_balance"], grand["mooe_expenses"], grand["mooe_balance"],
                grand["co_expenses"], grand["co_balance"]
            ))

            # Office Summary (Row 7 onwards)
            office_data = []
            for row in rows[6:]:
                if not row or not row[0]: continue
                office_data.append((
                    str(row[0]),
                    self._safe_float(row[1]),
                    self._safe_float(row[2]),
                    self._safe_float(row[3]),
                    self._safe_float(row[4]),
                    self._safe_float(row[5]),
                    self._safe_float(row[6])
                ))
            
            conn.execute("DELETE FROM office_summary")
            conn.executemany("""
                INSERT INTO office_summary (
                    office, budget_threshold, obligated, earnmarked, expenses, balance, utilization_pct
                ) VALUES (?,?,?,?,?,?,?)
            """, office_data)
            conn.commit()

    def _sync_expenditures(self, sheet):
        data = []
        rows = list(sheet.iter_rows(values_only=True))
        for row in rows[2:]: # Skip header
            if not any(row[1:]): continue
            data.append((
                str(row[0]) if row[0] else "",
                str(row[1]),
                str(row[2]) if row[2] else "",
                self._safe_float(row[3]),
                self._safe_float(row[4]),
                self._safe_float(row[5]),
                self._safe_float(row[6]),
                self._safe_float(row[7])
            ))
        
        with self._connect() as conn:
            conn.execute("DELETE FROM expenditures")
            conn.executemany("""
                INSERT INTO expenditures (
                    no, object_name, account_code, allotted_budget, earnmarked,
                    actual_obligated, expenses_total, balance_budget
                ) VALUES (?,?,?,?,?,?,?,?)
            """, data)
            conn.commit()

    def _sync_aip_programs(self, sheet):
        data = []
        rows = list(sheet.iter_rows(values_only=True))
        for i, row in enumerate(rows):
            if i < 8: continue # Skip headers
            if not any(row): continue
            
            is_header = 1 if row[0] is None and row[1] is None else 0
            
            data.append((
                str(row[0]) if row[0] else None,
                str(row[1]) if row[1] else None,
                str(row[2]) if row[2] else "",
                self._safe_float(row[3]),
                self._safe_float(row[4]),
                self._safe_float(row[5]),
                self._safe_float(row[6]),
                self._safe_float(row[7]),
                self._safe_float(row[8]),
                self._safe_float(row[9]),
                self._safe_float(row[10]),
                str(row[11]) if row[11] else None,
                self._safe_float(row[12]),
                str(row[13]) if row[13] else None,
                is_header
            ))
            
        with self._connect() as conn:
            conn.execute("DELETE FROM aip_programs")
            conn.executemany("""
                INSERT INTO aip_programs (
                    ref_code, office, description, ps_budget, mooe_budget, co_budget,
                    total_budget, earnmarked, obligated, expenses, balance,
                    status, utilization_pct, quarterly_schedule, is_header
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, data)
            conn.commit()

    def sync_remote(self):
        # Placeholder for remote CSV sync logic
        # Could use requests to get CSV and parse with csv module
        return False, "Remote sync not implemented yet (requires public CSV link)."

if __name__ == "__main__":
    sync = SpreadsheetSync("gad_budget.db", "GAD BUDGET MONITORING 2026.xlsx")
    ok, msg = sync.sync()
    print(msg)
