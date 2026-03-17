import openpyxl

def check_monitoring():
    wb = openpyxl.load_workbook("GAD BUDGET MONITORING 2026.xlsx", data_only=True)
    if "MONITORING" not in wb.sheetnames:
        print("MONITORING sheet not found.")
        return
    
    sheet = wb["MONITORING"]
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True)):
        print(f"Row {i+1}: {row}")

if __name__ == "__main__":
    check_monitoring()
